#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ROUTER PORFIN (596 / 575)
Lee los archivos exportados desde Porfin:
  - 596YYYYMMDD.CSV  → posiciones, valoraciones
  - SKCLI575*.CSV    → causaciones
Soporta: observaciones por fila, export Excel con openpyxl.
"""
from __future__ import annotations

import re
import json
import logging
from pathlib import Path
from typing import Optional, List, Dict, Any

import pandas as pd
import numpy as np
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
import io
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import get_data_dir

logger = logging.getLogger("porfin")
router = APIRouter()

UMBRAL_VAR_PCT = 5.0
UMBRAL_DIF_VAL = 1.0


# ── Utilidades ──────────────────────────────────────────────────────────────

def _base_dir() -> Path:
    return get_data_dir()


def _dirs(fecha: str = "") -> List[Path]:
    base = _base_dir()
    if fecha:
        return [base / fecha, base]
    return [base]


def _find_596(fecha: str = "") -> Optional[Path]:
    for d in _dirs(fecha):
        if not d.exists():
            continue
        for f in sorted(d.iterdir()):
            if not f.is_file():
                continue
            stem = f.stem.upper()
            if fecha and fecha not in stem and d == _base_dir():
                continue
            if re.match(r"^596\d{6,8}$", stem):
                return f
            if "596" in stem and f.suffix.upper() in (".CSV", ".TXT"):
                return f
    return None


def _find_575(fecha: str = "") -> Optional[Path]:
    fallback_generico = None
    for d in _dirs(fecha):
        if not d.exists():
            continue
        for f in sorted(d.iterdir()):
            if not f.is_file():
                continue
            stem = f.stem.upper()
            if fecha and d == _base_dir():
                compact = re.sub(r"\D", "", stem)
                if fecha not in compact:
                    if fallback_generico is None and "575" in stem and f.suffix.upper() in (".CSV", ".TXT"):
                        fallback_generico = f
                    continue
            if "575" in stem and f.suffix.upper() in (".CSV", ".TXT"):
                return f
            if re.match(r"^SKCLI", stem) and f.suffix.upper() in (".CSV", ".TXT"):
                return f
    return fallback_generico


def _read_porfin_csv(path: Path) -> pd.DataFrame:
    try:
        df = pd.read_csv(path, sep=";", dtype=str, encoding="latin-1", on_bad_lines="skip")
        df.columns = df.columns.str.strip()
        df = df[~df.iloc[:, 0].str.match(r"^-+$", na=True)]
        df = df[df.iloc[:, 0].notna() & df.iloc[:, 0].str.strip().ne("")]
        first_col = df.columns[0]
        df = df[df[first_col].str.strip() != first_col.strip()]
        return df.reset_index(drop=True)
    except Exception as e:
        logger.error(f"Error leyendo {path}: {e}")
        return pd.DataFrame()


def _normalize_num_col(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip().str.replace(",", "", regex=False)
    return pd.to_numeric(s, errors="coerce")


def _col(df: pd.DataFrame, *candidates) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
        for col in df.columns:
            if col.strip().upper() == c.upper():
                return col
        for col in df.columns:
            if c.upper() in col.upper():
                return col
    return None


def _cargar_596(fecha: str = "") -> pd.DataFrame:
    path = _find_596(fecha)
    if not path:
        return pd.DataFrame()
    df = _read_porfin_csv(path)
    if df.empty:
        return df

    renames = {
        _col(df, "Especie"):          "ESPECIE",
        _col(df, "Titulo"):           "TITULO",
        _col(df, "ISIN"):             "ISIN",
        _col(df, "Nemot", "Nemo"):    "NEMO",
        _col(df, "Emision"):          "EMISION",
        _col(df, "F.Vcto"):           "VCTO",
        _col(df, "Valor Nominal"):    "NOMINAL",
        _col(df, "Facial"):           "FACIAL",
        _col(df, "Mod"):              "METODO",
        _col(df, "F.Compra"):         "F_COMPRA",
        _col(df, "Valor Compra"):     "VALOR_COMPRA",
        _col(df, "Moned"):            "MONEDA",
        _col(df, "Valor Mercado Or"): "VLR_MER_OR",
        _col(df, "Valor Mercado"):    "VLR_MERCADO",
        _col(df, "T.Mer"):            "TASA_MER",
        _col(df, "Precio"):           "PRECIO",
        _col(df, "Tasa"):             "TASA",
        _col(df, "TIR"):              "TIR",
        _col(df, "Dur", "Duration"):  "DURACION",
        _col(df, "Llave"):            "LLAVE",
        _col(df, "Por"):              "PORTAFOLIO",
        _col(df, "PUC"):              "PUC",
        _col(df, "Cla"):              "CLASE",
        _col(df, "Fte"):              "FUENTE_TIT",
    }
    renames = {k: v for k, v in renames.items() if k and k != v}
    df = df.rename(columns=renames)

    for c in ["NOMINAL", "VALOR_COMPRA", "VLR_MERCADO", "VLR_MER_OR", "PRECIO", "TIR", "TASA_MER", "DURACION"]:
        if c in df.columns:
            df[c] = _normalize_num_col(df[c])

    if "CLASE" in df.columns:
        df = df[df["CLASE"].str.strip().isin(["Neg", "NoN"])]

    df["ARCHIVO"] = path.name
    df["_ROW_ID"] = df.index.astype(str)
    return df.reset_index(drop=True)


def _cargar_575(fecha: str = "") -> pd.DataFrame:
    path = _find_575(fecha)
    if not path:
        return pd.DataFrame()
    df = _read_porfin_csv(path)
    if df.empty:
        return df

    renames = {
        _col(df, "Especie"):                     "ESPECIE",
        _col(df, "Título", "Titulo"):            "TITULO",
        _col(df, "Inver"):                        "INVER",
        _col(df, "F.Vcto"):                       "VCTO",
        _col(df, "Vlr Nominal", "Valor Nominal"): "NOMINAL",
        _col(df, "Facial"):                       "FACIAL",
        _col(df, "Mod"):                          "METODO",
        _col(df, "Desde"):                        "DESDE",
        _col(df, "Hasta"):                        "HASTA",
        _col(df, "Vlr Mer. Ant"):                 "VLR_MER_ANT",
        _col(df, "Vlr Mer. Hoy"):                 "VLR_MER_HOY",
        _col(df, "Adeudados"):                    "ADEUDADOS",
        _col(df, "Causación Mer", "Causacion Mer"): "CAUSACION_MER",
        _col(df, "Causación TIR", "Causacion TIR"): "CAUSACION_TIR",
        _col(df, "ISIN"):                         "ISIN",
        _col(df, "Precio"):                       "PRECIO",
        _col(df, "TIR.Mercado", "TIR"):           "TIR",
        _col(df, "Moned"):                        "MONEDA",
        _col(df, "Mnd Val An"):                   "MND_VAL_ANT",
        _col(df, "Mnd Val"):                      "MND_VAL_HOY",
        _col(df, "Dif.cambio"):                   "DIF_CAMBIO",
        _col(df, "Causación Moneda", "Causacion Moneda"): "CAUSACION_MONEDA",
        _col(df, "Causación Tasa", "Causacion Tasa"):     "CAUSACION_TASA",
        _col(df, "Dias"):                         "DIAS",
        _col(df, "Por"):                          "PORTAFOLIO",
        _col(df, "Est"):                          "ESTADO",
    }
    renames = {k: v for k, v in renames.items() if k and k != v}
    df = df.rename(columns=renames)

    for c in ["NOMINAL", "VLR_MER_ANT", "VLR_MER_HOY", "CAUSACION_MER", "CAUSACION_TIR",
              "PRECIO", "TIR", "DIF_CAMBIO", "CAUSACION_MONEDA", "CAUSACION_TASA"]:
        if c in df.columns:
            df[c] = _normalize_num_col(df[c])

    if "ESPECIE" in df.columns:
        df = df[~df["ESPECIE"].str.strip().str.match(r"^-+$", na=True)]
        df = df[df["ESPECIE"].str.strip().ne("")]

    df["ARCHIVO"] = path.name
    df["_ROW_ID"] = df.index.astype(str)
    return df.reset_index(drop=True)


def _fechas_disponibles() -> List[str]:
    base = _base_dir()
    fechas = set()
    patron = re.compile(r"(\d{8})")
    if not base.exists():
        return []
    for d in base.iterdir():
        if d.is_dir() and re.match(r"^\d{8}$", d.name):
            fechas.add(d.name)
    for f in base.iterdir():
        if not f.is_file():
            continue
        stem = f.stem.upper()
        matches = patron.findall(stem)
        for fecha in matches:
            try:
                pd.to_datetime(fecha, format="%Y%m%d", errors="raise")
            except Exception:
                continue
            if stem.startswith("596") and stem == f"596{fecha}":
                fechas.add(fecha)
            elif "575" in stem and fecha in re.sub(r"\D", "", stem):
                fechas.add(fecha)
    return sorted(fechas)


# ── Observaciones ────────────────────────────────────────────────────────────

def _obs_file(fecha: str, modulo: str) -> Path:
    return _base_dir() / fecha / f"obs_{modulo}_{fecha}.json"


def _load_obs(fecha: str, modulo: str) -> Dict[str, str]:
    f = _obs_file(fecha, modulo)
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _save_obs(fecha: str, modulo: str, obs: Dict[str, str]) -> None:
    f = _obs_file(fecha, modulo)
    f.parent.mkdir(parents=True, exist_ok=True)
    f.write_text(json.dumps(obs, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/fechas")
def get_fechas():
    return {"fechas": _fechas_disponibles()}


@router.get("/resumen/{fecha}")
def resumen_porfin(fecha: str):
    df = _cargar_596(fecha)
    if df.empty:
        return {"error": f"No se encontró 596 para {fecha}", "fecha": fecha, "total": 0}

    total = len(df)
    tipo_dist = {}
    if "ESPECIE" in df.columns:
        df["_TIPO"] = df["ESPECIE"].str.strip().str.split().str[:2].str.join(" ")
        tipo_dist = df["_TIPO"].value_counts().head(20).to_dict()

    port_dist = df["PORTAFOLIO"].str.strip().value_counts().head(20).to_dict() if "PORTAFOLIO" in df.columns else {}
    mon_dist  = df["MONEDA"].str.strip().value_counts().to_dict() if "MONEDA" in df.columns else {}

    val_total = None
    if "VLR_MERCADO" in df.columns:
        val_total = round(float(df["VLR_MERCADO"].dropna().sum()), 2)

    df_575 = _cargar_575(fecha)
    caus_mer = round(float(df_575["CAUSACION_MER"].dropna().sum()), 2) if not df_575.empty and "CAUSACION_MER" in df_575.columns else None
    sin_precio = int(df["PRECIO"].isna().sum()) if "PRECIO" in df.columns else 0

    return {
        "fecha": fecha,
        "archivo_596": df["ARCHIVO"].iloc[0] if "ARCHIVO" in df.columns else "",
        "archivo_575": df_575["ARCHIVO"].iloc[0] if not df_575.empty and "ARCHIVO" in df_575.columns else "",
        "total_posiciones": total,
        "distribucion_especie": tipo_dist,
        "distribucion_portafolio": port_dist,
        "distribucion_moneda": mon_dist,
        "valor_mercado_total": val_total,
        "causacion_mercado_total": caus_mer,
        "sin_precio": sin_precio,
        "total_alertas": sin_precio,
    }


@router.get("/posiciones/{fecha}")
def posiciones_porfin(
    fecha: str,
    busqueda: Optional[str] = Query(None),
    moneda: Optional[str] = Query(None),
    portafolio: Optional[str] = Query(None),
    solo_alertas: bool = Query(False),
    limit: int = Query(2000),
):
    df = _cargar_596(fecha)
    if df.empty:
        return []

    obs_map = _load_obs(fecha, "porfin596")

    if "PRECIO" in df.columns:
        df["ALERTA"] = df["PRECIO"].isna()
    else:
        df["ALERTA"] = False

    if busqueda:
        mask = pd.Series(False, index=df.index)
        for c in ["ISIN", "NEMO", "ESPECIE", "LLAVE", "TITULO"]:
            if c in df.columns:
                mask = mask | df[c].astype(str).str.upper().str.contains(busqueda.upper(), na=False)
        df = df[mask]
    if moneda and "MONEDA" in df.columns:
        df = df[df["MONEDA"].str.upper().str.contains(moneda.upper(), na=False)]
    if portafolio and "PORTAFOLIO" in df.columns:
        df = df[df["PORTAFOLIO"].str.upper().str.contains(portafolio.upper(), na=False)]
    if solo_alertas:
        df = df[df["ALERTA"]]

    cols = [c for c in [
        "_ROW_ID", "CLASE", "FUENTE_TIT", "ESPECIE", "TITULO", "ISIN", "NEMO",
        "EMISION", "VCTO", "NOMINAL", "MONEDA", "VLR_MERCADO", "VLR_MER_OR",
        "PRECIO", "TIR", "TASA_MER", "DURACION", "PORTAFOLIO", "LLAVE", "PUC", "ALERTA"
    ] if c in df.columns]

    records = df[cols].replace({float("nan"): None, True: True, False: False}).head(limit).to_dict(orient="records")
    for r in records:
        r["OBS"] = obs_map.get(str(r.get("_ROW_ID", "")), "")
    return records


@router.get("/causaciones/{fecha}")
def causaciones_porfin(fecha: str, limit: int = Query(1000)):
    df = _cargar_575(fecha)
    if df.empty:
        return {"error": f"No se encontró 575 para {fecha}", "fecha": fecha}

    obs_map = _load_obs(fecha, "porfin575")

    caus_mer_total = round(float(df["CAUSACION_MER"].dropna().sum()), 2) if "CAUSACION_MER" in df.columns else None
    caus_tir_total = round(float(df["CAUSACION_TIR"].dropna().sum()), 2) if "CAUSACION_TIR" in df.columns else None

    if "CAUSACION_MER" in df.columns and "CAUSACION_TIR" in df.columns:
        df["DIF_MER_TIR"] = (df["CAUSACION_MER"] - df["CAUSACION_TIR"]).round(2)

    vlr_hoy = round(float(df["VLR_MER_HOY"].dropna().sum()), 2) if "VLR_MER_HOY" in df.columns else None
    vlr_ant = round(float(df["VLR_MER_ANT"].dropna().sum()), 2) if "VLR_MER_ANT" in df.columns else None

    detalle = df.replace({float("nan"): None}).head(limit).to_dict(orient="records")
    for r in detalle:
        r["OBS"] = obs_map.get(str(r.get("_ROW_ID", "")), "")

    return {
        "fecha": fecha,
        "archivo": df["ARCHIVO"].iloc[0] if "ARCHIVO" in df.columns else "",
        "total_registros": len(df),
        "causacion_mercado_total": caus_mer_total,
        "causacion_tir_total": caus_tir_total,
        "diferencia_mer_tir": round(caus_mer_total - caus_tir_total, 2) if caus_mer_total and caus_tir_total else None,
        "vlr_mercado_hoy": vlr_hoy,
        "vlr_mercado_ant": vlr_ant,
        "delta_valoracion": round(vlr_hoy - vlr_ant, 2) if vlr_hoy and vlr_ant else None,
        "detalle": detalle,
    }


@router.get("/errores/{fecha}")
def errores_porfin(fecha: str):
    df = _cargar_596(fecha)
    if df.empty:
        return {"error": f"No se encontró 596 para {fecha}", "total_errores": 0}

    errores: List[Dict] = []

    if "PRECIO" in df.columns:
        for _, r in df[df["PRECIO"].isna()].iterrows():
            errores.append({
                "tipo": "SIN PRECIO", "severidad": "ALTO",
                "isin": str(r.get("ISIN", "")).strip(),
                "especie": str(r.get("ESPECIE", "")).strip(),
                "portafolio": str(r.get("PORTAFOLIO", "")).strip(),
                "detalle": "Posición sin precio de valoración",
            })

    if "VLR_MERCADO" in df.columns:
        for _, r in df[df["VLR_MERCADO"] < 0].iterrows():
            errores.append({
                "tipo": "VALORACION NEGATIVA", "severidad": "ALTO",
                "isin": str(r.get("ISIN", "")).strip(),
                "especie": str(r.get("ESPECIE", "")).strip(),
                "portafolio": str(r.get("PORTAFOLIO", "")).strip(),
                "detalle": f"Vlr Mercado: {r.get('VLR_MERCADO', 0):,.2f}",
            })

    if "NOMINAL" in df.columns:
        for _, r in df[df["NOMINAL"].isna() | (df["NOMINAL"] == 0)].head(30).iterrows():
            errores.append({
                "tipo": "NOMINAL CERO", "severidad": "BAJO",
                "isin": str(r.get("ISIN", "")).strip(),
                "especie": str(r.get("ESPECIE", "")).strip(),
                "portafolio": str(r.get("PORTAFOLIO", "")).strip(),
                "detalle": "Nominal cero o nulo",
            })

    return {
        "fecha": fecha,
        "total_errores": len(errores),
        "resumen_severidad": {
            "ALTO":  sum(1 for e in errores if e["severidad"] == "ALTO"),
            "MEDIO": sum(1 for e in errores if e["severidad"] == "MEDIO"),
            "BAJO":  sum(1 for e in errores if e["severidad"] == "BAJO"),
        },
        "resumen_tipo": {t: sum(1 for e in errores if e["tipo"] == t) for t in set(e["tipo"] for e in errores)},
        "errores": errores[:500],
    }


@router.get("/portafolios/{fecha}")
def portafolios(fecha: str):
    df = _cargar_596(fecha)
    if df.empty or "PORTAFOLIO" not in df.columns:
        return []
    grp = df.groupby("PORTAFOLIO").agg(
        posiciones=("ISIN", "count"),
        valor_mercado=("VLR_MERCADO", "sum"),
    ).reset_index()
    grp["valor_mercado"] = grp["valor_mercado"].round(2)
    return grp.replace({float("nan"): None}).to_dict(orient="records")


@router.get("/monedas/{fecha}")
def monedas_596(fecha: str):
    df = _cargar_596(fecha)
    if df.empty or "MONEDA" not in df.columns:
        return []
    grp = df.groupby("MONEDA").agg(
        posiciones=("ISIN", "count"),
        valor_mercado=("VLR_MERCADO", "sum"),
    ).reset_index()
    grp["valor_mercado"] = grp["valor_mercado"].round(2)
    return grp.sort_values("valor_mercado", ascending=False).replace({float("nan"): None}).to_dict(orient="records")


@router.get("/variaciones/{fecha_inicio}/{fecha_fin}")
def variaciones_porfin(fecha_inicio: str, fecha_fin: str, umbral: float = Query(5.0)):
    df_i = _cargar_596(fecha_inicio)
    df_f = _cargar_596(fecha_fin)
    if df_i.empty or df_f.empty:
        return []
    col_id = next((c for c in ["ISIN", "LLAVE", "TITULO"] if c in df_i.columns and c in df_f.columns), None)
    if not col_id or "VLR_MERCADO" not in df_i.columns:
        return []
    df_i[col_id] = df_i[col_id].astype(str).str.strip()
    df_f[col_id] = df_f[col_id].astype(str).str.strip()
    merged = (
        df_i[[col_id, "VLR_MERCADO", "ESPECIE"]].rename(columns={"VLR_MERCADO": "VAL_I"})
        .merge(df_f[[col_id, "VLR_MERCADO"]].rename(columns={"VLR_MERCADO": "VAL_F"}), on=col_id)
    )
    merged = merged.dropna(subset=["VAL_I", "VAL_F"])
    merged = merged[merged["VAL_I"] != 0]
    merged["VAR_ABS"] = (merged["VAL_F"] - merged["VAL_I"]).round(2)
    merged["VAR_PCT"] = ((merged["VAR_ABS"] / merged["VAL_I"].abs()) * 100).round(4)
    merged["ANORMAL"] = merged["VAR_PCT"].abs() > umbral
    return merged.replace({float("nan"): None}).sort_values("VAR_PCT", key=abs, ascending=False).to_dict(orient="records")


# ── Observaciones ────────────────────────────────────────────────────────────

from pydantic import BaseModel as PydanticBase
from typing import Dict as TDict

class ObsPayload(PydanticBase):
    fila: str
    obs: str
    modulo: str = "porfin596"


@router.post("/observaciones/{fecha}")
def guardar_obs(fecha: str, body: ObsPayload):
    obs_map = _load_obs(fecha, body.modulo)
    if body.obs.strip():
        obs_map[body.fila] = body.obs.strip()
    else:
        obs_map.pop(body.fila, None)
    _save_obs(fecha, body.modulo, obs_map)
    return {"ok": True, "total_obs": len(obs_map)}


@router.get("/observaciones/{fecha}")
def get_obs(fecha: str, modulo: str = Query("porfin596")):
    return _load_obs(fecha, modulo)


# ── Export Excel ─────────────────────────────────────────────────────────────

@router.get("/excel/{fecha}")
def export_excel(fecha: str):
    """
    Genera un Excel de revisión diaria con:
    - Hoja 596: posiciones con observaciones, alertas resaltadas
    - Hoja 575: causaciones con alertas
    - Hoja Resumen: KPIs del día
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side,
            numbers as xl_numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado. Ejecuta: pip install openpyxl")

    df596 = _cargar_596(fecha)
    df575 = _cargar_575(fecha)
    obs596 = _load_obs(fecha, "porfin596")
    obs575 = _load_obs(fecha, "porfin575")

    wb = Workbook()

    # ── Estilos ──
    AZUL_HEADER  = PatternFill("solid", fgColor="1F3864")
    ROJO_ALERTA  = PatternFill("solid", fgColor="C00000")
    AMARILLO_OBS = PatternFill("solid", fgColor="FFD700")
    GRIS_ALT     = PatternFill("solid", fgColor="F2F2F2")
    VERDE        = PatternFill("solid", fgColor="E2EFDA")

    font_header = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_alerta = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    font_normal = Font(name="Calibri", size=9)
    font_titulo = Font(name="Calibri", bold=True, size=13, color="1F3864")

    alin_centro  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alin_izq     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    alin_der     = Alignment(horizontal="right",  vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _header_row(ws, cols, row_n=1):
        for j, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row_n, column=j, value=col_name)
            cell.fill = AZUL_HEADER
            cell.font = font_header
            cell.alignment = alin_centro
            cell.border = border

    def _auto_width(ws, max_w=40):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)

    def _fmt_num(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return None
        try:
            return float(v)
        except Exception:
            return v

    # ── Hoja Resumen ────
    ws_res = wb.active
    ws_res.title = "Resumen"

    ws_res.merge_cells("A1:F1")
    c = ws_res["A1"]
    c.value = f"REVISIÓN DIARIA PORFIN — {fecha[:4]}-{fecha[4:6]}-{fecha[6:]}"
    c.font = font_titulo
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 28

    kpis = [
        ("Posiciones 596", len(df596) if not df596.empty else 0),
        ("Sin Precio", int(df596["PRECIO"].isna().sum()) if not df596.empty and "PRECIO" in df596.columns else 0),
        ("Valor Mercado Total", _fmt_num(df596["VLR_MERCADO"].sum()) if not df596.empty and "VLR_MERCADO" in df596.columns else None),
        ("Causación Mercado (575)", _fmt_num(df575["CAUSACION_MER"].sum()) if not df575.empty and "CAUSACION_MER" in df575.columns else None),
        ("Registros 575", len(df575) if not df575.empty else 0),
        ("Observaciones 596", len(obs596)),
        ("Observaciones 575", len(obs575)),
        ("Archivo 596", df596["ARCHIVO"].iloc[0] if not df596.empty and "ARCHIVO" in df596.columns else "—"),
        ("Archivo 575", df575["ARCHIVO"].iloc[0] if not df575.empty and "ARCHIVO" in df575.columns else "—"),
    ]
    for i, (k, v) in enumerate(kpis, 3):
        ws_res.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws_res.cell(row=i, column=1).fill = GRIS_ALT if i % 2 else PatternFill()
        cell_v = ws_res.cell(row=i, column=2, value=v)
        cell_v.font = font_normal
        if isinstance(v, float) and abs(v) > 1_000:
            cell_v.number_format = '#,##0.00'
    ws_res.column_dimensions["A"].width = 30
    ws_res.column_dimensions["B"].width = 28

    # ── Hoja 596 ────
    ws596 = wb.create_sheet("Posiciones 596")
    if not df596.empty:
        cols596 = [c for c in [
            "CLASE", "ESPECIE", "TITULO", "ISIN", "NEMO", "EMISION", "VCTO",
            "NOMINAL", "MONEDA", "VLR_MERCADO", "VLR_MER_OR",
            "PRECIO", "TIR", "TASA_MER", "DURACION",
            "PORTAFOLIO", "LLAVE", "PUC", "ALERTA", "_ROW_ID"
        ] if c in df596.columns]

        # Agregar alertas
        if "PRECIO" in df596.columns:
            df596["ALERTA"] = df596["PRECIO"].isna()

        display_cols = [c for c in cols596 if c != "_ROW_ID"] + ["OBS"]
        _header_row(ws596, display_cols, 1)
        ws596.row_dimensions[1].height = 18
        ws596.freeze_panes = "A2"

        for i, (_, row) in enumerate(df596[cols596].iterrows(), 2):
            row_id = str(row.get("_ROW_ID", i - 2))
            obs_text = obs596.get(row_id, "")
            is_alerta = bool(row.get("ALERTA", False))
            is_obs = bool(obs_text)

            for j, col in enumerate(display_cols, 1):
                if col == "OBS":
                    val = obs_text
                else:
                    val = _fmt_num(row.get(col))

                cell = ws596.cell(row=i, column=j, value=val)
                cell.font = font_alerta if is_alerta else font_normal
                cell.border = border

                if is_alerta and col not in ("OBS",):
                    cell.fill = ROJO_ALERTA
                elif is_obs:
                    cell.fill = AMARILLO_OBS
                elif i % 2 == 0:
                    cell.fill = GRIS_ALT

                if col in ("VLR_MERCADO", "VLR_MER_OR", "NOMINAL", "VALOR_COMPRA"):
                    cell.number_format = '#,##0.00'
                    cell.alignment = alin_der
                elif col in ("PRECIO", "TIR", "TASA_MER", "DURACION"):
                    cell.number_format = '0.000000'
                    cell.alignment = alin_der
                else:
                    cell.alignment = alin_izq

        _auto_width(ws596)

    # ── Hoja 575 ────
    ws575 = wb.create_sheet("Causaciones 575")
    if not df575.empty:
        cols575 = [c for c in [
            "ESPECIE", "TITULO", "ISIN", "VCTO", "NOMINAL",
            "VLR_MER_ANT", "VLR_MER_HOY", "CAUSACION_MER", "CAUSACION_TIR",
            "DIF_MER_TIR", "PRECIO", "TIR", "MONEDA", "PORTAFOLIO", "_ROW_ID"
        ] if c in df575.columns]

        if "CAUSACION_MER" in df575.columns and "CAUSACION_TIR" in df575.columns:
            df575["DIF_MER_TIR"] = (df575["CAUSACION_MER"] - df575["CAUSACION_TIR"]).round(2)

        display_cols575 = [c for c in cols575 if c != "_ROW_ID"] + ["OBS"]
        _header_row(ws575, display_cols575, 1)
        ws575.row_dimensions[1].height = 18
        ws575.freeze_panes = "A2"

        for i, (_, row) in enumerate(df575[cols575].iterrows(), 2):
            row_id = str(row.get("_ROW_ID", i - 2))
            obs_text = obs575.get(row_id, "")
            dif = row.get("DIF_MER_TIR")
            is_alerta = dif is not None and not (isinstance(dif, float) and np.isnan(dif)) and abs(float(dif)) > UMBRAL_DIF_VAL

            for j, col in enumerate(display_cols575, 1):
                val = obs_text if col == "OBS" else _fmt_num(row.get(col))
                cell = ws575.cell(row=i, column=j, value=val)
                cell.font = font_alerta if is_alerta else font_normal
                cell.border = border

                if is_alerta:
                    cell.fill = ROJO_ALERTA
                elif obs_text:
                    cell.fill = AMARILLO_OBS
                elif i % 2 == 0:
                    cell.fill = GRIS_ALT

                if col in ("VLR_MER_ANT", "VLR_MER_HOY", "CAUSACION_MER", "CAUSACION_TIR", "DIF_MER_TIR", "NOMINAL"):
                    cell.number_format = '#,##0.00'
                    cell.alignment = alin_der
                else:
                    cell.alignment = alin_izq

        _auto_width(ws575)

    # ── Generar ────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"Revision_Porfin_{fecha}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Moneda normalizada ───────────────────────────────────────────────────────

def _norm_moneda(m: str) -> str:
    m = str(m).strip()
    if m.startswith("US"):  return "USD"
    if m in ("$", "COP"):   return "COP"
    if m.startswith("UU") or m.startswith("UVR"): return "UVR"
    if m.startswith("EU"):  return "EUR"
    if m.startswith("UK"):  return "GBP"
    if m.startswith("BRL"): return "BRL"
    if m.startswith("MXN"): return "MXN"
    if m.startswith("YEN"): return "JPY"
    if m.startswith("E") or m.startswith("EI") or m.startswith("EF"): return "FIC"
    return m


# ── Fondos (575) ──────────────────────────────────────────────────────────────

def _get_fondos_df(df575: pd.DataFrame) -> pd.DataFrame:
    """Agrega 575 por fondo (parte antes del guión en PORTAFOLIO)."""
    df = df575.copy()
    df["FONDO"] = df["PORTAFOLIO"].str.strip().str.split("-").str[0].str.strip()
    df["TIPO_PORT"] = df["PORTAFOLIO"].str.strip().str.split("-").str[1].str.strip().str[:1]
    df["MON_NORM"] = df["MONEDA"].apply(_norm_moneda)

    grp = df.groupby("FONDO").agg(
        portafolios=("PORTAFOLIO", "nunique"),
        posiciones=("ISIN", "count"),
        vlr_hoy=("VLR_MER_HOY", "sum"),
        vlr_ant=("VLR_MER_ANT", "sum"),
        caus_mer=("CAUSACION_MER", "sum"),
        caus_tir=("CAUSACION_TIR", "sum"),
    ).reset_index()
    grp["delta_valor"] = grp["vlr_hoy"] - grp["vlr_ant"]
    grp["dif_caus"] = grp["caus_mer"] - grp["caus_tir"]
    for c in ["vlr_hoy", "vlr_ant", "caus_mer", "caus_tir", "delta_valor", "dif_caus"]:
        grp[c] = grp[c].round(2)
    return grp.sort_values("vlr_hoy", ascending=False)


@router.get("/fondos/{fecha}")
def fondos_575(fecha: str):
    """Dashboard de fondos: agrupación por fondo, totales, causaciones, delta."""
    df = _cargar_575(fecha)
    if df.empty:
        return {"error": f"No se encontró 575 para {fecha}", "fondos": []}

    grp = _get_fondos_df(df)

    # Monedas por fondo
    df["FONDO"] = df["PORTAFOLIO"].str.strip().str.split("-").str[0].str.strip()
    df["MON_NORM"] = df["MONEDA"].apply(_norm_moneda)
    monedas_por_fondo: Dict[str, Dict] = {}
    for fondo, sub in df.groupby("FONDO"):
        mon_grp = sub.groupby("MON_NORM").agg(
            posiciones=("ISIN", "count"),
            vlr_hoy=("VLR_MER_HOY", "sum"),
            caus_mer=("CAUSACION_MER", "sum"),
        ).reset_index()
        monedas_por_fondo[fondo] = mon_grp.round(2).to_dict(orient="records")

    # Totales globales
    total_vlr = float(df["VLR_MER_HOY"].sum())
    total_caus = float(df["CAUSACION_MER"].sum())

    return {
        "fecha": fecha,
        "total_vlr_mercado": round(total_vlr, 2),
        "total_causacion_mer": round(total_caus, 2),
        "total_fondos": len(grp),
        "fondos": grp.replace({float("nan"): None}).to_dict(orient="records"),
        "monedas_por_fondo": monedas_por_fondo,
    }


@router.get("/fondos/{fecha}/{fondo}")
def fondo_detalle(
    fecha: str,
    fondo: str,
    moneda: Optional[str] = Query(None),
    solo_alertas: bool = Query(False),
):
    """Detalle de posiciones de un fondo específico (575)."""
    df = _cargar_575(fecha)
    if df.empty:
        return []
    obs_map = _load_obs(fecha, "porfin575")

    df["FONDO"] = df["PORTAFOLIO"].str.strip().str.split("-").str[0].str.strip()
    df["MON_NORM"] = df["MONEDA"].apply(_norm_moneda)
    df = df[df["FONDO"].str.upper() == fondo.upper()]

    if moneda:
        df = df[df["MON_NORM"].str.upper() == moneda.upper()]

    if "CAUSACION_MER" in df.columns and "CAUSACION_TIR" in df.columns:
        df["DIF_MER_TIR"] = (df["CAUSACION_MER"] - df["CAUSACION_TIR"]).round(2)
        df["ALERTA_CAUS"] = df["DIF_MER_TIR"].abs() > UMBRAL_DIF_VAL

    if solo_alertas and "ALERTA_CAUS" in df.columns:
        df = df[df["ALERTA_CAUS"]]

    cols = [c for c in [
        "_ROW_ID", "PORTAFOLIO", "ESPECIE", "TITULO", "ISIN", "VCTO",
        "NOMINAL", "MON_NORM", "MONEDA",
        "VLR_MER_ANT", "VLR_MER_HOY", "CAUSACION_MER", "CAUSACION_TIR",
        "DIF_MER_TIR", "ALERTA_CAUS", "PRECIO", "TIR", "METODO", "ESTADO"
    ] if c in df.columns]

    records = df[cols].replace({float("nan"): None, True: True, False: False}).to_dict(orient="records")
    for r in records:
        r["OBS"] = obs_map.get(str(r.get("_ROW_ID", "")), "")
    return records


@router.get("/monedas_fondos/{fecha}")
def monedas_fondos(fecha: str):
    """Distribución por moneda normalizada en 575."""
    df = _cargar_575(fecha)
    if df.empty:
        return []
    df["MON_NORM"] = df["MONEDA"].apply(_norm_moneda)
    grp = df.groupby("MON_NORM").agg(
        posiciones=("ISIN", "count"),
        vlr_hoy=("VLR_MER_HOY", "sum"),
        vlr_ant=("VLR_MER_ANT", "sum"),
        caus_mer=("CAUSACION_MER", "sum"),
        caus_tir=("CAUSACION_TIR", "sum"),
    ).reset_index()
    grp["delta"] = grp["vlr_hoy"] - grp["vlr_ant"]
    for c in ["vlr_hoy", "vlr_ant", "caus_mer", "caus_tir", "delta"]:
        grp[c] = grp[c].round(2)
    return grp.sort_values("vlr_hoy", ascending=False).replace({float("nan"): None}).to_dict(orient="records")


# ── Alertas TIR / DV (596) ────────────────────────────────────────────────────

_MET_COL_596 = "Mét"   # columna 'Mét' en 596 — método de curva: TC, QSI, MC*

def _met_col_596(df: pd.DataFrame) -> Optional[str]:
    """Encuentra la columna 'Mét' (método de curva) en el 596."""
    for c in df.columns:
        # buscar la col que contiene TC, QSI, MC
        if df[c].astype(str).str.strip().isin(["TC", "QSI", "MC3-I", "MC4-E", "MC1-I", "PS"]).any():
            return c
    return None


def _fecha_siguiente(fecha: str) -> str:
    """Retorna fecha + 1 día en formato YYYYMMDD."""
    try:
        from datetime import datetime, timedelta
        d = datetime.strptime(fecha, "%Y%m%d") + timedelta(days=1)
        return d.strftime("%Y%m%d")
    except Exception:
        return ""


@router.get("/alertas_tir/{fecha}")
def alertas_tir_596(fecha: str, fecha_emision_hoy: Optional[str] = Query(None)):
    """
    Detecta posiciones en el 596 que están valorando a TIR/curva:
    - FUENTE_TIT = 1DV → al vencimiento explícito (siempre alerta)
    - Mét = TC (Tasa Cupón = precio a TIR implícita) — TES/CDTs sin precio de mercado
    - METODO DV pero VCTO != mañana ni hoy → título al vencimiento fuera de plazo normal
    Excluye DV con VCTO == fecha o fecha+1 (normal: vencen hoy/mañana).
    """
    df = _cargar_596(fecha)
    if df.empty:
        return {"error": f"No se encontró 596 para {fecha}", "alertas": []}

    fecha_emision = fecha_emision_hoy or fecha
    fecha_sig = _fecha_siguiente(fecha)
    alertas: List[Dict] = []
    met_c = _met_col_596(df)

    for _, r in df.iterrows():
        motivos = []
        metodo = str(r.get("METODO", "")).strip()
        fuente = str(r.get("FUENTE_TIT", "")).strip()
        emision = str(r.get("EMISION", "")).strip()
        vcto = str(r.get("VCTO", "")).strip()

        # Excluir emitidos hoy (normal que usen TIR el primer día)
        if emision == fecha_emision:
            continue

        # 1DV = fuente al vencimiento (siempre es alerta si no vence inminente)
        if fuente == "1DV":
            motivos.append("FUENTE 1DV (al vencimiento)")

        # DV = método al vencimiento — solo alerta si NO vence hoy ni mañana
        if metodo == "DV" and vcto not in (fecha, fecha_sig, ""):
            motivos.append(f"METODO DV (vcto {vcto})")

        # Mét = TC (Tasa Cupón = valoración a TIR implícita sin precio de mercado)
        if met_c and str(r.get(met_c, "")).strip() == "TC":
            motivos.append("Mét TC (Tasa Cupón/TIR)")

        if motivos:
            alertas.append({
                "motivo":    " | ".join(motivos),
                "isin":      str(r.get("ISIN", "")).strip(),
                "especie":   str(r.get("ESPECIE", "")).strip(),
                "titulo":    str(r.get("TITULO", "")).strip(),
                "emision":   emision,
                "vcto":      vcto,
                "metodo":    metodo,
                "fuente_tit": fuente,
                "met_curva": str(r.get(met_c, "")).strip() if met_c else "",
                "precio":    None if pd.isna(r.get("PRECIO")) else r.get("PRECIO"),
                "tir":       None if pd.isna(r.get("TIR")) else r.get("TIR"),
                "portafolio": str(r.get("PORTAFOLIO", "")).strip(),
                "llave":     str(r.get("LLAVE", "")).strip(),
                "nominal":   None if pd.isna(r.get("NOMINAL")) else r.get("NOMINAL"),
                "vlr_mercado": None if pd.isna(r.get("VLR_MERCADO")) else r.get("VLR_MERCADO"),
            })

    resumen: Dict[str, int] = {}
    for a in alertas:
        for parte in a["motivo"].split(" | "):
            # Normalizar la clave (quitar el vcto variable del DV)
            key = parte if "METODO DV" not in parte else "METODO DV (al vencimiento)"
            resumen[key] = resumen.get(key, 0) + 1

    return {
        "fecha": fecha,
        "total_alertas": len(alertas),
        "resumen_motivo": resumen,
        "alertas": alertas,
    }


@router.get("/alertas_tir_575/{fecha}")
def alertas_tir_575(fecha: str):
    """
    Detecta en el 575 posiciones valorando a TIR cuando no corresponde.
    Usa columna Mét (QES-SI = curva, MC* = matriz, TC sería alerta).
    """
    df = _cargar_575(fecha)
    if df.empty:
        return {"error": f"No se encontró 575 para {fecha}", "alertas": []}

    alertas: List[Dict] = []

    # Buscar col Mét en 575
    met_c = None
    for c in df.columns:
        if df[c].astype(str).str.strip().isin(["QES-SI", "MC3-I", "MC4-E", "MC1-I", "QSI"]).any():
            met_c = c
            break

    for _, r in df.iterrows():
        motivos = []
        metodo = str(r.get("METODO", "")).strip()

        if metodo == "DV":
            motivos.append("METODO DV")

        # Si en 575 hay una columna Mét con TC
        if met_c and str(r.get(met_c, "")).strip() in ("TC", "TC   "):
            motivos.append("Mét TC (Tasa Cupón)")

        if motivos:
            alertas.append({
                "motivo":     " | ".join(motivos),
                "isin":       str(r.get("ISIN", "")).strip(),
                "especie":    str(r.get("ESPECIE", "")).strip(),
                "titulo":     str(r.get("TITULO", "")).strip(),
                "metodo":     metodo,
                "met_curva":  str(r.get(met_c, "")).strip() if met_c else "",
                "precio":     r.get("PRECIO"),
                "tir":        r.get("TIR"),
                "vlr_mer_hoy": r.get("VLR_MER_HOY"),
                "causacion_mer": r.get("CAUSACION_MER"),
                "portafolio": str(r.get("PORTAFOLIO", "")).strip(),
            })

    return {
        "fecha": fecha,
        "total_alertas": len(alertas),
        "alertas": alertas,
    }


# ── Config de columnas para alerta TIR ────────────────────────────────────────
# Permite al usuario ver qué columna y qué valor indica TIR en cada archivo

@router.get("/config_columnas/{fecha}")
def config_columnas(fecha: str):
    """
    Devuelve el mapa de columnas relevantes detectadas en 596 y 575
    para la configuración de alertas TIR/DV.
    """
    df596 = _cargar_596(fecha)
    df575 = _cargar_575(fecha)

    met_c_596 = _met_col_596(df596) if not df596.empty else None
    met_c_575 = None
    if not df575.empty:
        for c in df575.columns:
            if df575[c].astype(str).str.strip().isin(["QES-SI", "MC3-I", "MC4-E", "QSI"]).any():
                met_c_575 = c
                break

    return {
        "596": {
            "col_metodo":   "METODO",
            "col_fuente":   "FUENTE_TIT",
            "col_met_curva": met_c_596,
            "metodo_dv_valor": "DV",
            "fuente_dv_valor": "1DV",
            "met_curva_tc_valor": "TC",
            "metodo_valores_unicos": sorted(df596["METODO"].dropna().unique().tolist()) if not df596.empty and "METODO" in df596.columns else [],
            "fuente_valores_unicos": sorted(df596["FUENTE_TIT"].dropna().unique().tolist()) if not df596.empty and "FUENTE_TIT" in df596.columns else [],
            "met_curva_valores_unicos": sorted(df596[met_c_596].dropna().str.strip().unique().tolist()) if met_c_596 and not df596.empty else [],
        },
        "575": {
            "col_metodo":   "METODO",
            "col_met_curva": met_c_575,
            "metodo_dv_valor": "DV",
            "metodo_valores_unicos": sorted(df575["METODO"].dropna().unique().tolist()) if not df575.empty and "METODO" in df575.columns else [],
            "met_curva_valores_unicos": sorted(df575[met_c_575].dropna().str.strip().unique().tolist()) if met_c_575 and not df575.empty else [],
        },
    }


# ── Motor de verificación de valoración ───────────────────────────────────────
#
# Replica la lógica del notebook MVALORACIONTOTAL para calcular
# VALORACION_MANUAL y CAUSACION_MANUAL y compararlas contra Porfin.
#
# Fórmulas por TIPO (del notebook):
#  FCPE / DFI / FONDOS DE PENSION → vlr_mer_or * moneda
#  FM                              → vlr_mer_or * moneda * precio
#  FIC / FCP                       → vlr_mer_or * precio
#  CASH / CTA AHORROS / COLATERAL  → nominal * moneda
#  ADR / ETF / ACCION INTERNACIONAL→ nominal * moneda * precio
#  BONO INTERNACIONAL y similares  → nominal * moneda * (precio / 100)
#  ACCION / ESTRATEGIAS            → nominal * precio
#  BONO / CDT / TES PESOS y simil  → nominal * (precio / 100)
#  FONDOCRENA                      → vlr_mer_or * precio
#  ANTICIPO / FIDECOMISO / LOTE…   → vlr_mercado (de Porfin, sin recalcular)
#  default                         → vlr_mercado

_TIPOS_PORCENTAJE = {
    "BONO", "CDT", "STRUCTURADO", "SUBORDINADO", "TES PESOS",
    "TITULARIZADORA", "PAPEL COMERCIAL",
}
_TIPOS_PORC_MONEDA = {
    "BONO INTERNACIONAL", "BONO UVR", "CDT UVR", "SUBORDINADO INTERNACIONAL",
    "SUBORDINADO UVR", "TES UVR", "TIPS", "TREASURY", "YANKEE", "TBILL",
    "TD", "NESTR",
}

# Factores especiales por LLAVE (del legacy Revisiones_Valoracion).
# Cada entrada puede ser:
#   float  → factor para vlr_mer_or * moneda * factor  (fondos JZ*/JH*)
#   dict   → {"factor": ..., "tipo": "nominal"|"fijo"|"vlr_mer_or"}
_SPECIAL_LLAVES: Dict[str, Any] = {
    # Fondos especiales con factor numérico
    "JZCJ3": 1.064920,  "JZAXO": 1.673050,  "JZAVN": 1.082659,
    "JHCKT": 0.938460,  "JGDKT": 1.160131,  "JZAKT": 1.076858,
    "JHPKT": 0.958241,  "JZAPJ": 0.816438,  "JZDP6": 0.845205,
    "JHM0Y": 0.978898,  "JZAQK": 1.041950,  "JZAOK": 1.183056,
    "JHHJ2": 0.899836,  "JHIJ4": 1.082530,  "JZA1X": 1.024280,
    "JZBJQ": 1.000409,  "JHDJQ": 0.946999,  "JZANQ": 0.984023,
    "JZEXO": 1.638932,  "JZAKE": 0.943544,  "JZAY0": 1.173927,
    "JZASS": 1.008909,  "JHKK0": 0.976514,  "JZAVX": 0.981900,
    "JHS6":  0.885150,  "JZAVW": 1.080053,  "JZFVW": 1.079829,
    # Fondos con nominal fijo (factor = nominal_de_referencia)
    "2ACAC": {"factor": 180138.96,         "tipo": "nominal"},
    "2FIU1": {"factor": 9687.34,           "tipo": "nominal"},
    "8LTGK": {"factor": 1054949999.63,     "tipo": "fijo"},
    "2BVHA": {"factor": 18368.206568,      "tipo": "vlr_mer_or"},
    "FVBQA": {"factor": 13593.114169,      "tipo": "vlr_mer_or"},
}

# Factores moneda para conversión a COP  (cargado en runtime desde monedas)
def _factor_moneda_estatico(m: str) -> float:
    """Valor fijo 1 si no hay tabla de tasas. La API de verificación
    carga la tabla real y pasa el factor real."""
    m = str(m).strip().upper()
    if m in ("COP", "$", "COP$", "COL$", "PESOS", "PESO", ""):
        return 1.0
    return 1.0  # sin tabla de tasas, no convierte — flag de "pendiente"


def _calcular_valor_fila(
    tipo: str,
    nominal: float,
    vlr_mer_or: float,
    vlr_mercado: float,
    precio: float,
    moneda_factor: float,
    llave: str = "",
) -> Optional[float]:
    """
    Calcula el valor de mercado manual para una posición.
    Devuelve None si faltan datos esenciales.
    """
    tipo  = str(tipo).strip().upper()
    llave = str(llave).strip().upper()

    # SPECIAL_LLAVES: fondos con factor propio
    if llave and llave in _SPECIAL_LLAVES:
        spec = _SPECIAL_LLAVES[llave]
        if isinstance(spec, float):
            # JZ*/JH* → vlr_mer_or * moneda * factor
            if pd.notna(vlr_mer_or):
                return vlr_mer_or * moneda_factor * spec
            return None
        else:
            t = spec.get("tipo", "")
            f = spec.get("factor", 1.0)
            if t == "nominal":
                return f * moneda_factor if pd.notna(f) else None
            elif t == "fijo":
                return f * moneda_factor if pd.notna(f) else None
            elif t == "vlr_mer_or":
                return vlr_mer_or * f if pd.notna(vlr_mer_or) else None

    if tipo in ("FCPE", "DFI", "FONDOS DE PENSION"):
        if pd.notna(vlr_mer_or):
            return vlr_mer_or * moneda_factor
    elif tipo == "FM":
        if pd.notna(vlr_mer_or) and pd.notna(precio):
            return vlr_mer_or * moneda_factor * precio
    elif tipo in ("FIC", "FCP"):
        if pd.notna(vlr_mer_or) and pd.notna(precio):
            return vlr_mer_or * precio
    elif tipo in ("CASH", "CTA AHORROS", "COLATERAL"):
        if pd.notna(nominal):
            return nominal * moneda_factor
    elif tipo in ("ADR", "ETF", "ACCION INTERNACIONAL"):
        if pd.notna(nominal) and pd.notna(precio):
            return nominal * moneda_factor * precio
    elif tipo in _TIPOS_PORC_MONEDA:
        if pd.notna(nominal) and pd.notna(precio):
            return nominal * moneda_factor * (precio / 100.0)
    elif tipo in ("ACCION", "ESTRATEGIAS"):
        if pd.notna(nominal) and pd.notna(precio):
            return nominal * precio
    elif tipo in _TIPOS_PORCENTAJE:
        if pd.notna(nominal) and pd.notna(precio):
            return nominal * (precio / 100.0)
    elif tipo == "FONDOCRENA":
        if pd.notna(vlr_mer_or) and pd.notna(precio):
            return vlr_mer_or * precio
    elif tipo in ("ANTICIPO", "FIDEICOMISO", "LOTE", "CREDITO", "FORWARD"):
        return vlr_mercado if pd.notna(vlr_mercado) else None
    else:
        return vlr_mercado if pd.notna(vlr_mercado) else None
    return None


def _cargar_especies(cfg: Dict) -> Dict[str, str]:
    """Carga el archivo Especies.csv y retorna dict LLAVE → TIPO."""
    path = Path(cfg.get("ref_especies", ""))
    if not path.exists():
        return {}
    try:
        df = pd.read_csv(path, sep=";", encoding="latin-1", dtype=str,
                         on_bad_lines="skip")
        df.columns = df.columns.astype(str).str.strip()
        llave_col = next((c for c in df.columns if c.upper() == "LLAVE"), None)
        tipo_col  = next((c for c in df.columns if c.upper() == "TIPO"), None)
        if not llave_col or not tipo_col:
            return {}
        df = df[[llave_col, tipo_col]].dropna()
        df[llave_col] = df[llave_col].astype(str).str.strip().str.upper()
        df[tipo_col]  = df[tipo_col].astype(str).str.strip().str.upper()
        df = df[df[llave_col] != ""].drop_duplicates(llave_col, keep="first")
        return df.set_index(llave_col)[tipo_col].to_dict()
    except Exception as e:
        logger.warning(f"No se pudo cargar Especies.csv: {e}")
        return {}


def _cargar_precios_insumos(fecha: str) -> Dict[str, float]:
    """
    Carga precios del día desde los proveedores Infovalmer disponibles.
    Prioridad: SP → SW → MX → NOTAS → MX_RV
    Retorna dict ISIN/NEMO (upper) → precio.
    """
    try:
        import sys
        sys.path.insert(0, str(Path(__file__).parent))
        from insumos import cargar_sp, cargar_sw, cargar_mx, cargar_notas, cargar_mx_rv
    except ImportError:
        return {}

    precios: Dict[str, float] = {}
    for cargador in [cargar_mx_rv, cargar_notas, cargar_mx, cargar_sw, cargar_sp]:
        try:
            df = cargador(fecha)
            if df.empty or "PRECIO" not in df.columns:
                continue
            for _, r in df.iterrows():
                pid = str(r.get("ID", r.get("ISIN", r.get("NEMO", "")))).strip().upper()
                p   = float(r["PRECIO"]) if pd.notna(r["PRECIO"]) else None
                if pid and p is not None and pid not in precios:
                    precios[pid] = p
                    # También indexar por NEMO si distinto
                    nemo = str(r.get("NEMO", "")).strip().upper()
                    if nemo and nemo != pid and nemo not in precios:
                        precios[nemo] = p
        except Exception as e:
            logger.warning(f"Error cargando precios insumos: {e}")
    return precios


def _cargar_tasas_cambio(fecha: str) -> Dict[str, float]:
    """Retorna dict MONEDA_NORM → factor_a_COP."""
    try:
        from insumos import cargar_monedas
        df = cargar_monedas(fecha)
        if df.empty:
            return {}
        tasas: Dict[str, float] = {}
        for _, r in df.iterrows():
            moneda = str(r.get("MONEDA", "")).strip().upper()
            tasa   = r.get("TASA_COP")
            if moneda and tasa and pd.notna(tasa):
                tasas[moneda] = float(tasa)
        return tasas
    except Exception:
        return {}


def _resolver_moneda_factor(moneda_str: str, tasas: Dict[str, float]) -> float:
    m = str(moneda_str).strip().upper()
    if m in ("COP", "$", "COP$", "COL$", "PESOS", "PESO", ""):
        return 1.0
    if m.startswith("US"):
        return tasas.get("USD", 1.0)
    if m.startswith("EU"):
        return tasas.get("EUR", 1.0)
    if m.startswith("UVR"):
        return tasas.get("UVR", 1.0)
    if m.startswith("UU"):
        return tasas.get("UVR", 1.0)
    if m.startswith("UK") or m == "GBP":
        return tasas.get("GBP", 1.0)
    if m.startswith("BRL"):
        return tasas.get("BRL", 1.0)
    if m.startswith("MXN"):
        return tasas.get("MXN", 1.0)
    if m in ("E", "EI", "EF", "FIC"):
        return 1.0
    return tasas.get(m, 1.0)


@router.get("/verificacion/{fecha}")
def verificacion_valoracion(
    fecha: str,
    umbral_pct: float = Query(1.0, description="% diferencia para marcar como alerta"),
    umbral_abs: float = Query(1000.0, description="Diferencia absoluta para alerta"),
    solo_alertas: bool = Query(False),
    portafolio: Optional[str] = Query(None),
):
    """
    Verifica la valoración de Porfin contra el cálculo manual con precios
    de proveedores Infovalmer (SP/SW/MX/NOTAS).
    Cada fila tiene: VALORACION_PORFIN, PRECIO_INFOVALMER, TIPO,
    VALORACION_MANUAL, DIF_ABS, DIF_PCT, ALERTA, MOTIVO_ALERTA.
    """
    from config import get_config
    cfg = get_config()

    df = _cargar_596(fecha)
    if df.empty:
        return {"error": f"No se encontró 596 para {fecha}", "filas": []}

    mapa_tipo   = _cargar_especies(cfg)
    precios_inf = _cargar_precios_insumos(fecha)
    tasas       = _cargar_tasas_cambio(fecha)

    def _norm(s):
        return str(s).strip().upper()

    def _keya(especie: str, isin: str, nemo: str, moneda: str, llave: str) -> str:
        """Determina la clave de búsqueda de precio según el tipo de especie (KEYA logic del legacy)."""
        esp = especie.upper()
        # Fondos de inversión colectiva y similares: usar MONEDA como clave
        if any(k in esp for k in ("FIC", "FCP", "FCPE", "CC", "DFI")):
            return moneda.strip().upper()
        # Acciones preferentes o ordinarias locales: usar NEMO
        if any(k in esp for k in ("APR", "AOR", "ACCION")):
            return nemo
        # CASH: clave literal
        if "CASH" in esp:
            return "CASH"
        # ADR: usar ISIN
        if "ADR" in esp:
            return isin
        # Todo lo demás (CDT, bonos, TES, ETF, Yankee…): usar ISIN como clave principal
        return isin if isin else nemo

    resultados = []
    for _, r in df.iterrows():
        llave  = _norm(r.get("LLAVE", ""))
        isin   = _norm(r.get("ISIN",  ""))
        nemo   = _norm(r.get("NEMO",  ""))
        moneda = str(r.get("MONEDA", "")).strip()
        especie = str(r.get("ESPECIE", "")).strip()

        # Tipo de activo
        tipo = (mapa_tipo.get(llave) or
                mapa_tipo.get(isin)  or
                mapa_tipo.get(nemo)  or "")

        # KEYA: clave de lookup de precio según tipo de especie
        keya = _keya(especie, isin, nemo, moneda, llave)

        # Precio Infovalmer: buscar por KEYA primero, luego ISIN → NEMO → LLAVE
        precio_inf = (precios_inf.get(keya) or
                      precios_inf.get(isin)  or
                      precios_inf.get(nemo)  or
                      precios_inf.get(llave))

        # Factor moneda
        mfactor = _resolver_moneda_factor(moneda, tasas)

        nominal     = r.get("NOMINAL")
        vlr_mer_or  = r.get("VLR_MER_OR")
        vlr_mercado = r.get("VLR_MERCADO")

        val_porfin = vlr_mercado  # lo que dice Porfin

        # Calcular valoración manual (usa precio Infovalmer si disponible,
        # si no usa el precio del propio 596)
        precio_para_calculo = precio_inf if precio_inf is not None else r.get("PRECIO")

        val_manual = _calcular_valor_fila(
            tipo=tipo,
            nominal=float(nominal) if pd.notna(nominal) else float("nan"),
            vlr_mer_or=float(vlr_mer_or) if pd.notna(vlr_mer_or) else float("nan"),
            vlr_mercado=float(vlr_mercado) if pd.notna(vlr_mercado) else float("nan"),
            precio=float(precio_para_calculo) if pd.notna(precio_para_calculo) else float("nan"),
            moneda_factor=mfactor,
            llave=llave,
        )

        # Diferencias
        dif_abs = None
        dif_pct = None
        alerta  = False
        motivos = []

        if val_manual is not None and val_porfin is not None and pd.notna(val_porfin) and pd.notna(val_manual):
            dif_abs = round(val_porfin - val_manual, 2)
            if val_manual != 0:
                dif_pct = round((dif_abs / abs(val_manual)) * 100, 4)
            if abs(dif_abs) > umbral_abs:
                alerta = True
                motivos.append(f"Dif abs {dif_abs:,.2f}")
            if dif_pct is not None and abs(dif_pct) > umbral_pct:
                alerta = True
                motivos.append(f"Dif % {dif_pct:.2f}%")
        elif val_manual is None and pd.notna(val_porfin):
            if tipo == "":
                motivos.append("Tipo activo desconocido (Especies.csv no cargado o LLAVE sin mapeo)")
            elif precio_inf is None:
                motivos.append(f"Sin precio Infovalmer — tipo {tipo}")
                alerta = True

        if solo_alertas and not alerta:
            continue
        if portafolio and "PORTAFOLIO" in df.columns:
            if portafolio.upper() not in _norm(r.get("PORTAFOLIO", "")):
                continue

        fila = {
            "LLAVE":              llave,
            "ISIN":               isin,
            "NEMO":               nemo,
            "KEYA":               keya,
            "ESPECIE":            especie,
            "TITULO":             str(r.get("TITULO",  "")).strip(),
            "TIPO":               tipo,
            "PORTAFOLIO":         str(r.get("PORTAFOLIO", "")).strip(),
            "MONEDA":             moneda,
            "NOMINAL":            None if pd.isna(nominal) else round(float(nominal), 2),
            "VLR_MER_OR":         None if pd.isna(vlr_mer_or) else round(float(vlr_mer_or), 2),
            "PRECIO_PORFIN":      None if pd.isna(r.get("PRECIO")) else round(float(r.get("PRECIO")), 6),
            "PRECIO_INFOVALMER":  round(precio_inf, 6) if precio_inf is not None else None,
            "FACTOR_MONEDA":      round(mfactor, 6),
            "VALORACION_PORFIN":  None if pd.isna(val_porfin) else round(float(val_porfin), 2),
            "VALORACION_MANUAL":  round(val_manual, 2) if val_manual is not None else None,
            "DIF_ABS":            dif_abs,
            "DIF_PCT":            dif_pct,
            "ALERTA":             alerta,
            "MOTIVO_ALERTA":      " | ".join(motivos) if motivos else "",
        }
        resultados.append(fila)

    total         = len(resultados)
    alertas_count = sum(1 for f in resultados if f["ALERTA"])
    sin_precio    = sum(1 for f in resultados if f["PRECIO_INFOVALMER"] is None)
    sin_tipo      = sum(1 for f in resultados if not f["TIPO"])

    return {
        "fecha": fecha,
        "total": total,
        "con_precio_infovalmer": total - sin_precio,
        "sin_precio_infovalmer": sin_precio,
        "sin_tipo": sin_tipo,
        "total_alertas": alertas_count,
        "umbral_pct_usado": umbral_pct,
        "umbral_abs_usado": umbral_abs,
        "filas": resultados,
    }


@router.get("/verificacion_causacion/{fecha}")
def verificacion_causacion(
    fecha: str,
    fecha_anterior: Optional[str] = Query(None, description="Fecha anterior YYYYMMDD — si no se indica, usa la más reciente disponible"),
    umbral_abs: float = Query(1000.0),
    solo_alertas: bool = Query(False),
    portafolio: Optional[str] = Query(None),
):
    """
    Verifica la causación del 575 contra el cálculo manual:
    CAUSACION_MANUAL = VLR_MER_HOY - VLR_MER_ANT  (método mercado)
    DIF_CAUSACION    = CAUSACION_MER (Porfin) - CAUSACION_MANUAL
    """
    df575 = _cargar_575(fecha)
    if df575.empty:
        return {"error": f"No se encontró 575 para {fecha}", "filas": []}

    # Fecha anterior
    if not fecha_anterior:
        fechas = sorted(set(
            d.name for d in _base_dir().iterdir()
            if d.is_dir() and re.match(r"^\d{8}$", d.name) and d.name < fecha
        ))
        fecha_anterior = fechas[-1] if fechas else None

    # Cargar 583 para INT_DIV (intereses y dividendos cobrados)
    df583    = _cargar_583(fecha)
    int_div  = _calcular_int_div(df583)    # dict CONSEC → suma cobros
    inc_ret  = _calcular_inc_ret_capital(df583)  # dict CONSEC → neto inc/ret

    # Tasas de cambio para CASH causación (NOMINAL * (MONEDA_T - MONEDA_Y))
    tasas_hoy = _cargar_tasas_cambio(fecha)
    tasas_ant = _cargar_tasas_cambio(fecha_anterior) if fecha_anterior else {}

    resultados = []
    for _, r in df575.iterrows():
        vlr_hoy  = r.get("VLR_MER_HOY")
        vlr_ant  = r.get("VLR_MER_ANT")
        caus_mer = r.get("CAUSACION_MER")
        caus_tir = r.get("CAUSACION_TIR")
        especie  = str(r.get("ESPECIE", "")).strip().upper()
        titulo   = str(r.get("TITULO",  "")).strip()
        isin     = str(r.get("ISIN",    "")).strip().upper()
        moneda   = str(r.get("MONEDA",  "")).strip()
        nominal  = r.get("NOMINAL")
        facial   = r.get("FACIAL")

        # Clave para INT_DIV: 575 usa TITULO como CONSEC
        consec_key = titulo if titulo else isin

        # Causación manual según TIPO
        caus_manual = None
        metodo_caus = "VLR_DIF"

        if "CASH" in especie:
            # CASH: NOMINAL * (MONEDA_T - MONEDA_Y) + VAL_X_TASA
            # VAL_X_TASA ≈ CAUSACION_MONEDA del propio 575 (Porfin ya lo calcula)
            mfactor_hoy = _resolver_moneda_factor(moneda, tasas_hoy)
            mfactor_ant = _resolver_moneda_factor(moneda, tasas_ant)
            caus_moneda = r.get("CAUSACION_MONEDA")
            if pd.notna(nominal) and mfactor_hoy != mfactor_ant:
                val_x_tasa = float(caus_moneda) if pd.notna(caus_moneda) else 0.0
                caus_manual = round(float(nominal) * (mfactor_hoy - mfactor_ant) + val_x_tasa, 2)
                metodo_caus = "CASH"
            elif pd.notna(vlr_hoy) and pd.notna(vlr_ant):
                caus_manual = round(float(vlr_hoy) - float(vlr_ant), 2)
        elif "CTA AHORROS" in especie:
            # CTA AHORROS: NOMINAL * (FACIAL/100/365)
            if pd.notna(nominal) and pd.notna(facial) and float(facial) != 0:
                caus_manual = round(float(nominal) * (float(facial) / 100.0 / 365.0), 2)
                metodo_caus = "CTA_AHORROS"
            elif pd.notna(vlr_hoy) and pd.notna(vlr_ant):
                caus_manual = round(float(vlr_hoy) - float(vlr_ant), 2)
        else:
            # General: VLR_MERCADO_T - VLR_MERCADO_Y + INT_DIV
            if pd.notna(vlr_hoy) and pd.notna(vlr_ant):
                int_div_val = int_div.get(consec_key, 0.0)
                caus_manual = round(float(vlr_hoy) - float(vlr_ant) + int_div_val, 2)
                metodo_caus = "VLR_DIF+INT_DIV" if int_div_val != 0 else "VLR_DIF"

        dif_abs = None
        alerta  = False
        motivos = []

        if caus_manual is not None and caus_mer is not None and pd.notna(caus_mer):
            dif_abs = round(float(caus_mer) - caus_manual, 2)
            if abs(dif_abs) > umbral_abs:
                alerta = True
                motivos.append(f"Dif causación {dif_abs:,.2f}")
        elif caus_mer is None or pd.isna(caus_mer):
            motivos.append("Sin causación Porfin")

        if solo_alertas and not alerta:
            continue
        if portafolio and "PORTAFOLIO" in df575.columns:
            if portafolio.upper() not in str(r.get("PORTAFOLIO", "")).upper():
                continue

        int_div_fila = int_div.get(consec_key, 0.0)

        resultados.append({
            "ESPECIE":              especie,
            "TITULO":               titulo,
            "ISIN":                 isin,
            "PORTAFOLIO":           str(r.get("PORTAFOLIO", "")).strip(),
            "MONEDA":               moneda,
            "NOMINAL":              None if pd.isna(nominal) else round(float(nominal), 2),
            "VLR_MER_ANT":          None if pd.isna(vlr_ant)  else round(float(vlr_ant),  2),
            "VLR_MER_HOY":          None if pd.isna(vlr_hoy)  else round(float(vlr_hoy),  2),
            "INT_DIV":              round(int_div_fila, 2) if int_div_fila else None,
            "CAUSACION_MER_PORFIN": None if pd.isna(caus_mer) else round(float(caus_mer), 2),
            "CAUSACION_TIR_PORFIN": None if pd.isna(caus_tir) else round(float(caus_tir), 2),
            "CAUSACION_MANUAL":     caus_manual,
            "METODO_CAUS":          metodo_caus,
            "DIF_CAUSACION":        dif_abs,
            "ALERTA":               alerta,
            "MOTIVO":               " | ".join(motivos) if motivos else "",
        })

    total         = len(resultados)
    alertas_count = sum(1 for f in resultados if f["ALERTA"])
    difs          = [f["DIF_CAUSACION"] for f in resultados if f["DIF_CAUSACION"] is not None]
    dif_total     = round(sum(difs), 2) if difs else 0.0

    return {
        "fecha":               fecha,
        "fecha_anterior":      fecha_anterior,
        "total":               total,
        "total_alertas":       alertas_count,
        "dif_total_causacion": dif_total,
        "umbral_abs_usado":    umbral_abs,
        "tiene_583":           not df583.empty,
        "total_int_div":       round(sum(int_div.values()), 2) if int_div else 0.0,
        "filas":               resultados,
    }


# ── Operaciones 583 ──────────────────────────────────────────────────────────
# Archivo 583{yyyymmdd}.csv = Informe de Operaciones (Porfin)
# Contiene: compras, ventas, cobro intereses, cobro dividendos,
#           incremento/retiro capital, etc.
# Se usa para calcular INT_DIV y ajustes de causación manual.

def _find_583(fecha: str = "") -> Optional[Path]:
    for d in _dirs(fecha):
        if not d.exists():
            continue
        for f in sorted(d.iterdir()):
            if not f.is_file():
                continue
            stem = f.stem.upper()
            if "583" in stem and f.suffix.upper() in (".CSV", ".TXT"):
                return f
    return None


def _cargar_583(fecha: str = "") -> pd.DataFrame:
    path = _find_583(fecha)
    if not path:
        return pd.DataFrame()
    df = _read_porfin_csv(path)
    if df.empty:
        return df

    renames = {
        _col(df, "Especie"):                      "ESPECIE",
        _col(df, "Transacc", "Transaccion"):      "TRANSACCION",
        _col(df, "Tipo Oper", "TipoOper"):        "TIPO_OPER",
        _col(df, "Consec"):                       "CONSEC",
        _col(df, "Moned"):                        "MONEDA",
        _col(df, "Val.Operac", "ValOperacion"):   "VAL_OPERACION",
        _col(df, "Precio Ope", "PrecioOpe"):      "PRECIO_OPE",
        _col(df, "Vr.Recibido", "VrRecibido"):    "VR_RECIBIDO",
        _col(df, "Por"):                          "PORTAFOLIO",
        _col(df, "ISIN"):                         "ISIN",
    }
    renames = {k: v for k, v in renames.items() if k and k != v}
    df = df.rename(columns=renames)

    for c in ["VAL_OPERACION", "PRECIO_OPE", "VR_RECIBIDO"]:
        if c in df.columns:
            df[c] = _normalize_num_col(df[c])

    df["ARCHIVO"] = path.name
    df["_ROW_ID"] = df.index.astype(str)
    return df.reset_index(drop=True)


def _calcular_int_div(df583: pd.DataFrame) -> Dict[str, float]:
    """Agrega intereses y dividendos cobrados por CONSEC (TITULO).
    Retorna dict CONSEC → suma VR_RECIBIDO de cobros de interés/dividendo."""
    if df583.empty or "TRANSACCION" not in df583.columns:
        return {}
    mask = df583["TRANSACCION"].str.strip().str.upper().isin(
        ["COBRO DIVIDENDOS", "COBRO INTERESES", "REINTEGRO CAPITAL"]
    )
    sub = df583[mask].copy()
    if sub.empty or "CONSEC" not in sub.columns:
        return {}
    sub["VR_RECIBIDO"] = pd.to_numeric(sub.get("VR_RECIBIDO", pd.Series()), errors="coerce").fillna(0)
    return sub.groupby("CONSEC")["VR_RECIBIDO"].sum().to_dict()


def _calcular_inc_ret_capital(df583: pd.DataFrame) -> Dict[str, float]:
    """Incrementos y retiros de capital netos por CONSEC."""
    if df583.empty or "TRANSACCION" not in df583.columns:
        return {}
    mask = df583["TRANSACCION"].str.strip().str.upper().isin(
        ["RETIRO CAPITAL", "INCREMENTO CAPITAL"]
    )
    sub = df583[mask].copy()
    if sub.empty or "CONSEC" not in sub.columns:
        return {}
    sub["VR_RECIBIDO"] = pd.to_numeric(sub.get("VR_RECIBIDO", pd.Series()), errors="coerce").fillna(0)
    sub["_NETO"] = sub.apply(
        lambda r: -r["VR_RECIBIDO"] if "RETIRO" in str(r["TRANSACCION"]).upper()
                  else r["VR_RECIBIDO"], axis=1
    )
    return sub.groupby("CONSEC")["_NETO"].sum().to_dict()


@router.get("/operaciones/{fecha}")
def operaciones_583(
    fecha: str,
    transaccion: Optional[str] = Query(None, description="Filtrar por tipo de transacción"),
    portafolio: Optional[str] = Query(None),
    limit: int = Query(1000),
):
    """
    Informe de Operaciones 583: compras, ventas, cobros de interés/dividendo,
    incremento/retiro capital. Base para ajustes de causación manual.
    """
    df = _cargar_583(fecha)
    if df.empty:
        return {"error": f"No se encontró 583 para {fecha}", "fecha": fecha, "registros": []}

    if transaccion and "TRANSACCION" in df.columns:
        df = df[df["TRANSACCION"].str.upper().str.contains(transaccion.upper(), na=False)]
    if portafolio and "PORTAFOLIO" in df.columns:
        df = df[df["PORTAFOLIO"].str.upper().str.contains(portafolio.upper(), na=False)]

    # Resúmenes
    tipos = df["TRANSACCION"].value_counts().to_dict() if "TRANSACCION" in df.columns else {}
    int_div = _calcular_int_div(df)
    inc_ret = _calcular_inc_ret_capital(df)

    total_int_div = round(sum(int_div.values()), 2) if int_div else 0.0
    total_inc_ret = round(sum(inc_ret.values()), 2) if inc_ret else 0.0

    cols = [c for c in [
        "_ROW_ID", "TRANSACCION", "TIPO_OPER", "CONSEC", "ESPECIE",
        "ISIN", "MONEDA", "VAL_OPERACION", "PRECIO_OPE", "VR_RECIBIDO", "PORTAFOLIO"
    ] if c in df.columns]

    return {
        "fecha":              fecha,
        "archivo":            df["ARCHIVO"].iloc[0] if "ARCHIVO" in df.columns else "",
        "total_registros":    len(df),
        "tipos_transaccion":  tipos,
        "total_int_div":      total_int_div,
        "total_inc_ret":      total_inc_ret,
        "registros":          df[cols].replace({float("nan"): None}).head(limit).to_dict(orient="records"),
    }


# ── Referencia para mitra.py (compatibilidad) ─────────────────────────────────
def _buscar_archivo_575(fecha: str) -> Optional[Path]:
    return _find_575(fecha)


def _leer_csv_robusto(path: Path) -> pd.DataFrame:
    return _read_porfin_csv(path)
